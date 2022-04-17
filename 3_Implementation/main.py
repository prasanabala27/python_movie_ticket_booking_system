import openpyxl
wb = openpyxl.load_workbook("mov.xlsx")
sheet = wb.active

# pylint: disable=missing-module-docstring
# pylint: disable=missing-class-docstring
# pylint: disable=missing-function-docstring
# pylint: disable=line-too-long
import userlogin
class BookTicket:
    def __init__(self):
        self.totalincome = 0
        self.name = ''
        self.age = ''
        self.phone =''
        self.price = ''
        self.user_row = ''
        self.user_seat = ''
        self.gender = ''
        self.rows = 10
        self.seats = 10
        self.l = []
        self.userlist = []
        self.currentincome = 0
        self.noticketpurchased = 0
        for i in range(self.rows):
            lin1 = ['S'] * self.seats
            lin2 = [{}] * self.seats
            self.l.append(lin1)
            self.userlist.append(lin2)
        self.menu()
    def show_seats(self):
        count = 1
        print("\n\t\t\tSEATS\n")
        for n in range(1, self.seats + 1):
            if n == 1:
                print('\t%d' % (n), end='\t')
            else:
                print(n, end='\t')
        print('\n')
        for i in self.l:
            print(count, end='\t')
            count += 1
            for j in i:
                print(j, end='\t')
            print('\n')

    def user(self):
        self.name = input('\nEnter your name:')
        self.gender = input('Enter your gender:')
        self.age = int(input('Enter your age:'))
        self.phone = int(input('Enter your phone number:'))
    def available(self, arow, brow):
        if arow <= self.rows and brow <= self.seats:
            return True
    def buy(self):
        row = sheet.max_row
        self.user_row = int(input('\nEnter the row of your seat 1-10:'))
        self.user_seat = int(input('Enter the seat number 1-10:'))
        if self.available(self.user_row, self.user_seat):
            print('\nRow', self.user_row, ' seat number:', self.user_seat)
            # self.priceofticket(self.user_row)
            print("Cost of the seat is Rs.", self.priceofticket(self.user_row))
            answer = input('\nChoose \'Yes\' or \'No\' to proceed :')
            if self.l[self.user_row - 1][self.user_seat - 1] == 'S':
                if (answer == 'Yes' or answer == 'yes'):
                    self.user()
                    self.l[self.user_row - 1][self.user_seat - 1] = 'B'
                    self.userlist[self.user_row - 1][self.user_seat - 1] = {'name': self.name, 'gender': self.gender,'age': self.age, 'phone': self.phone}
                    self.currentincome += self.priceofticket(self.user_row)
                    self.noticketpurchased += 1
                for i in range(1, row + 1):
                    cell_obj = sheet.cell(row=i, column=1)
                    print( cell_obj.value)
                ain = input("select the option: ")
                cell1 = sheet.cell(row=int(ain),column=1)
                print(cell1.value)
                for i in range(1, row + 1):
                    cell_obj = sheet.cell(row=i, column=2)
                    print( cell_obj.value)
                binn = input("enter the choice")
                cell2 = sheet.cell(row=int(binn),column=2)
                print(cell2.value)
                print("select the movie")
                for i in range(1, row + 1):
                    cell_obj = sheet.cell(row=i, column=3)
                    print( cell_obj.value)
                cin = input("enter the choice")
                cell3=sheet.cell(row=int(cin),column=3)
                print(cell3.value)
                print("select the movie")
                for i in range(1, row + 1):
                    cell_obj = sheet.cell(row=i, column=4)
                    print(cell_obj.value)
                din = input("enter the choice")
                cell4 = sheet.cell(row=int(din), column=4)
                print(cell4.value)
                print("Your Ticket is booked")              
                print(" CITY: COIMBATORE \n CINEMA HALL" + cell1.value +"\n MOVIE:" +cell2.value +"\n SCREEN:"+cell3.value+"\n TIME:"+ cell4.value)
                print("USER DETAILS:")
                print("Name:"+self.name)
            else:
                print("The seat is already booked")
        else:
            print("\n!!!!!!Seat doesn't exist!!!!!!")
    def priceofticket(self, a1):
        if self.rows * self.seats <= 60:
            self.price = 150
        else:
            if a1 <= (self.rows // 2):
                self.price = 120
            else:
                self.price = 100
        return self.price
    def seat_info(self):
        ainn = int(input("\nEnter the row no.:"))
        bin1 = int(input("Enter the seat no.:"))
        if self.available(ainn, bin1):
            if self.l[ainn - 1][bin1 - 1] == 'B':
                c = self.userlist[ainn - 1][bin1 - 1]
                print('Name:', c['name'])
                print('Gender:', c['gender'])
                print('Age:', c['age'])
                print('Phone no.:', c['phone'])
            else:
                print("The seat is empty/available")
        else:
            print("\n!!!!!!Seat doesn't exist!!!!!!")
    def customer(self):
        print('\n1.Show List of seats\n2.Book Ticket\n3.seat_info\n4.menu\n5.Exit')
        choice = int(input("Enter the choice:"))
        if choice == 1:
            self.show_seats()
            self.menu()
        elif choice == 2:
            self.buy()
            self.menu()
        elif choice == 3:
            self.seat_info()
            self.menu()
        elif choice == 4:
            self.menu()
        elif choice == 5:
            return None
    def menu(self):
        print('\n1.Show the seats\n2.Buy a ticket\n3.Statistics\n4.Show booked ticket user info\n0.Exit')
        choice = int(input("Enter the choice:"))
        if choice == 1:
            self.show_seats()
            self.menu()
        elif choice == 2:
            self.buy()
            self.menu()
        elif choice == 3:
            self.statistics()
            self.menu()
        elif choice == 4:
            self.seat_info()
            self.menu()
        elif choice == 0:
            return None
    def admin(self):
        print("adminmenu")
        print('\n1.seat_info\n2.menu\n3.Exit')
        choice = int(input("Enter the choice:"))
        if choice == 1:
            self.seat_info()
            self.menu()
        elif choice == 2:
            self.menu()
        elif choice == 3:
            return None
    def menu(self):
        print('\n1.Admin\n2.User\n3.Exit')
        choice = int(input("Enter the choice:"))
        if choice == 1:
            self.adminlog()
        elif choice == 2:
            self.customer()
            self.menu()
        elif choice == 3:
            return None
    def adminlog(self):
        print("Enter the admin id and password")
        adm = input("****admin****\n")
        paw = input("****password****\n")
        if adm=="admin" or paw=="password":
            self.admin()
        else:
            print("Enter correct id and password")
            self.menu()
ticketbooking = BookTicket()
